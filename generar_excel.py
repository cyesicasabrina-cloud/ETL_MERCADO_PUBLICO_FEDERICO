#!/usr/bin/env python3
"""
Script para generar archivos Excel a partir de los CSVs generados por el sistema ETL.
Convierte todos los archivos CSV en archivos Excel con formato y múltiples hojas.
"""

import os
import pandas as pd
from datetime import datetime
from pathlib import Path
import logging

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def crear_archivo_excel_completo():
    """Crea un archivo Excel completo con todas las hojas de datos"""
    
    # Rutas de archivos
    fecha = datetime.now().strftime("%Y%m%d")
    base_dir = Path("data")
    
    archivos_csv = {
        "Datos_Raw": base_dir / "raw" / f"licitaciones_estado_activas_raw_{fecha}.csv",
        "Datos_Clean": base_dir / "clean" / f"licitaciones_estado_activas_clean_{fecha}.csv",
        "Datos_Negocio": base_dir / "clean" / f"licitaciones_estado_activas_requested_{fecha}.csv"
    }
    
    # Archivo Excel de salida
    archivo_excel = base_dir / f"licitaciones_completo_{fecha}.xlsx"
    
    logger.info(f"📊 Creando archivo Excel completo: {archivo_excel}")
    
    try:
        with pd.ExcelWriter(archivo_excel, engine='openpyxl') as writer:
            
            # Procesar cada archivo CSV
            for nombre_hoja, ruta_csv in archivos_csv.items():
                if ruta_csv.exists():
                    logger.info(f"📋 Procesando {nombre_hoja}: {ruta_csv}")
                    
                    # Leer CSV
                    df = pd.read_csv(ruta_csv, encoding='utf-8-sig')
                    
                    # Escribir a Excel con formato
                    df.to_excel(
                        writer, 
                        sheet_name=nombre_hoja, 
                        index=False,
                        startrow=1
                    )
                    
                    # Obtener la hoja para aplicar formato
                    worksheet = writer.sheets[nombre_hoja]
                    
                    # Aplicar formato a los headers
                    from openpyxl.styles import Font, PatternFill, Alignment
                    
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    
                    # Formatear primera fila (headers)
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Ajustar ancho de columnas
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    logger.info(f"✅ Hoja '{nombre_hoja}' creada con {len(df)} filas y {len(df.columns)} columnas")
                    
                else:
                    logger.warning(f"⚠️  Archivo no encontrado: {ruta_csv}")
            
            # Crear hoja de resumen
            crear_hoja_resumen(writer, archivos_csv)
        
        logger.info(f"🎉 Archivo Excel completo creado: {archivo_excel}")
        return archivo_excel
        
    except Exception as e:
        logger.error(f"❌ Error creando archivo Excel: {e}")
        raise

def crear_hoja_resumen(writer, archivos_csv):
    """Crea una hoja de resumen con estadísticas"""
    
    logger.info("📊 Creando hoja de resumen...")
    
    # Recopilar estadísticas
    estadisticas = []
    
    for nombre_hoja, ruta_csv in archivos_csv.items():
        if ruta_csv.exists():
            try:
                df = pd.read_csv(ruta_csv, encoding='utf-8-sig')
                estadisticas.append({
                    'Hoja': nombre_hoja,
                    'Archivo': ruta_csv.name,
                    'Registros': len(df),
                    'Columnas': len(df.columns),
                    'Tamaño_KB': round(ruta_csv.stat().st_size / 1024, 2)
                })
            except Exception as e:
                logger.warning(f"Error procesando {ruta_csv}: {e}")
    
    if estadisticas:
        df_resumen = pd.DataFrame(estadisticas)
        df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
        
        # Formatear hoja de resumen
        worksheet = writer.sheets['Resumen']
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Formatear headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("✅ Hoja de resumen creada")

def crear_archivos_excel_individuales():
    """Crea archivos Excel individuales para cada CSV"""
    
    fecha = datetime.now().strftime("%Y%m%d")
    base_dir = Path("data")
    
    archivos_csv = {
        "raw": base_dir / "raw" / f"licitaciones_estado_activas_raw_{fecha}.csv",
        "clean": base_dir / "clean" / f"licitaciones_estado_activas_clean_{fecha}.csv",
        "negocio": base_dir / "clean" / f"licitaciones_estado_activas_requested_{fecha}.csv"
    }
    
    archivos_creados = []
    
    for tipo, ruta_csv in archivos_csv.items():
        if ruta_csv.exists():
            try:
                # Nombre del archivo Excel
                archivo_excel = base_dir / f"licitaciones_{tipo}_{fecha}.xlsx"
                
                logger.info(f"📊 Creando archivo Excel individual: {archivo_excel}")
                
                # Leer CSV
                df = pd.read_csv(ruta_csv, encoding='utf-8-sig')
                
                # Crear Excel con formato
                with pd.ExcelWriter(archivo_excel, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=f'Licitaciones_{tipo.title()}', index=False)
                    
                    # Formatear
                    worksheet = writer.sheets[f'Licitaciones_{tipo.title()}']
                    from openpyxl.styles import Font, PatternFill, Alignment
                    
                    # Formatear headers
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Ajustar ancho de columnas
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                archivos_creados.append(archivo_excel)
                logger.info(f"✅ Archivo Excel creado: {archivo_excel} ({len(df)} filas)")
                
            except Exception as e:
                logger.error(f"❌ Error creando Excel para {tipo}: {e}")
        else:
            logger.warning(f"⚠️  Archivo CSV no encontrado: {ruta_csv}")
    
    return archivos_creados

def mostrar_estadisticas_archivos():
    """Muestra estadísticas de los archivos generados"""
    
    logger.info("📊 ESTADÍSTICAS DE ARCHIVOS GENERADOS")
    logger.info("=" * 60)
    
    fecha = datetime.now().strftime("%Y%m%d")
    base_dir = Path("data")
    
    archivos = {
        "CSV Raw": base_dir / "raw" / f"licitaciones_estado_activas_raw_{fecha}.csv",
        "CSV Clean": base_dir / "clean" / f"licitaciones_estado_activas_clean_{fecha}.csv",
        "CSV Negocio": base_dir / "clean" / f"licitaciones_estado_activas_requested_{fecha}.csv",
        "Excel Completo": base_dir / f"licitaciones_completo_{fecha}.xlsx",
        "Excel Raw": base_dir / f"licitaciones_raw_{fecha}.xlsx",
        "Excel Clean": base_dir / f"licitaciones_clean_{fecha}.xlsx",
        "Excel Negocio": base_dir / f"licitaciones_negocio_{fecha}.xlsx"
    }
    
    total_size = 0
    
    for nombre, ruta in archivos.items():
        if ruta.exists():
            size_kb = round(ruta.stat().st_size / 1024, 2)
            total_size += size_kb
            
            if ruta.suffix == '.csv':
                try:
                    df = pd.read_csv(ruta, encoding='utf-8-sig')
                    logger.info(f"✅ {nombre}: {len(df)} filas, {len(df.columns)} columnas, {size_kb} KB")
                except:
                    logger.info(f"✅ {nombre}: {size_kb} KB (error leyendo CSV)")
            else:
                logger.info(f"✅ {nombre}: {size_kb} KB")
        else:
            logger.info(f"❌ {nombre}: No encontrado")
    
    logger.info(f"📊 Tamaño total: {total_size:.2f} KB ({total_size/1024:.2f} MB)")

def main():
    """Función principal"""
    
    logger.info("🚀 GENERADOR DE ARCHIVOS EXCEL")
    logger.info("=" * 50)
    
    try:
        # Crear archivo Excel completo
        archivo_completo = crear_archivo_excel_completo()
        
        # Crear archivos Excel individuales
        archivos_individuales = crear_archivos_excel_individuales()
        
        # Mostrar estadísticas
        mostrar_estadisticas_archivos()
        
        logger.info("\n🎉 PROCESO COMPLETADO EXITOSAMENTE")
        logger.info("=" * 50)
        logger.info(f"📊 Archivo Excel completo: {archivo_completo}")
        logger.info(f"📊 Archivos Excel individuales: {len(archivos_individuales)}")
        
        print(f"\n✅ Archivos Excel generados exitosamente!")
        print(f"📁 Archivo completo: {archivo_completo}")
        print(f"📁 Archivos individuales: {len(archivos_individuales)}")
        
        return 0
        
    except Exception as e:
        logger.error(f"❌ Error en el proceso: {e}")
        return 1

if __name__ == "__main__":
    exit_code = main()
    exit(exit_code)
